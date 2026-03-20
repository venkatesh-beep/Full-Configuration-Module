import io
from typing import Any

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from modules.ui_helpers import module_header, section_header


UPLOAD_SHEET_NAME = "Accrual_Policies_Upload"
EXISTING_POLICIES_SHEET = "Existing_Accrual_Policies"
EXISTING_ACCRUALS_SHEET = "Accruals_Reference"
ACCRUAL_REFERENCE_HOST = "https://saas-beeforce.labour.tech"
POLICY_TEMPLATE_COLUMNS = [
    "id", "name", "description", "accrualId", "grantType", "grantFrequency",
    "grantStartDate", "forceAvail", "grantExpiration", "grantExpiredAfter",
    "carryoverAmountMax", "prioritizeCarryoverBalance", "carryoverEncashmentAmountMax",
    "terminationEncashmentAmountMax", "manualEncashmentAmountMax",
    "startDate(DD/MM)1", "sendDate(DD/MM)1", "amount1",
    "startDate(DD/MM)2", "sendDate(DD/MM)2", "amount2",
    "grantStart1", "grantEnd1", "grantMax1", "grantAmount1",
    "grantStart2", "grantEnd2", "grantMax2", "grantAmount2",
    "ruleCondition1", "ruleValue1",
    "takingPaycodeID1", "takingAmount1", "takingPaycodeID2", "takingAmount2",
]

MAX_PRORATIONS = 50
MAX_GRANT_AMOUNTS = 20
MAX_RULES = 20
MAX_PAYCODES = 20


def _auth_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {st.session_state.token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def _bool_from_cell(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() == "true"


def _int_from_cell(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _float_from_cell(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _fetch_json(url: str, headers: dict[str, str]) -> tuple[list[dict[str, Any]], str | None]:
    try:
        response = requests.get(url, headers=headers, timeout=30)
    except requests.RequestException as exc:
        return [], str(exc)

    if response.status_code != 200:
        return [], f"HTTP {response.status_code}: {response.text[:200]}"

    try:
        data = response.json()
    except ValueError as exc:
        return [], f"Invalid JSON response: {exc}"

    if isinstance(data, list):
        return data, None
    return [], "Unexpected response format"


def _build_existing_policy_rows(policies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for policy in policies:
        row = {
            "id": policy.get("id"),
            "name": policy.get("name"),
            "description": policy.get("description"),
            "accrualId": (policy.get("accrual") or {}).get("id"),
            "grantType": policy.get("grantType"),
            "grantFrequency": policy.get("grantFrequency"),
            "grantStartDate": policy.get("grantStartDate"),
            "forceAvail": policy.get("forceAvail"),
            "grantExpiration": policy.get("grantExpiration"),
            "grantExpiredAfter": policy.get("grantExpiredAfter"),
            "carryoverAmountMax": policy.get("carryoverAmountMax"),
            "prioritizeCarryoverBalance": policy.get("prioritizeCarryoverBalance"),
            "carryoverEncashmentAmountMax": policy.get("carryoverEncashmentAmountMax"),
            "terminationEncashmentAmountMax": policy.get("terminationEncashmentAmountMax"),
            "manualEncashmentAmountMax": policy.get("manualEncashmentAmountMax"),
        }

        for index, item in enumerate(policy.get("grantProrations", []), start=1):
            row[f"startDate(DD/MM){index}"] = item.get("startDate")
            row[f"sendDate(DD/MM){index}"] = item.get("endDate")
            row[f"amount{index}"] = item.get("amount")

        for index, item in enumerate(policy.get("grantAmounts", []), start=1):
            row[f"grantStart{index}"] = item.get("start")
            row[f"grantEnd{index}"] = item.get("end")
            row[f"grantMax{index}"] = item.get("max")
            row[f"grantAmount{index}"] = item.get("amount")

        for index, item in enumerate(policy.get("grantAmountRules", []), start=1):
            row[f"ruleCondition{index}"] = item.get("condition")
            row[f"ruleValue{index}"] = item.get("value")

        for index, item in enumerate(policy.get("paycodes", []), start=1):
            row[f"takingPaycodeID{index}"] = ((item.get("paycode") or {}).get("id"))
            row[f"takingAmount{index}"] = item.get("amount")

        rows.append(row)

    return rows


def _build_template_workbook(
    policy_rows: list[dict[str, Any]],
    accrual_rows: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = UPLOAD_SHEET_NAME
    ws.append(POLICY_TEMPLATE_COLUMNS)
    ws.append([
        "",
        "TEST",
        "TEST",
        57,
        "FIXED",
        "MONTHLY",
        "2026-01-31",
        "FALSE",
        "DAY",
        1,
        "FALSE",
        "FALSE",
        "FALSE",
        "FALSE",
        "FALSE",
        "01/01",
        "01/31",
        2.5,
        "02/01",
        "02/25",
        2.5,
        0,
        5,
        "FALSE",
        1,
        6,
        "",
        "FALSE",
        2,
        "${T} >=0",
        2.5,
        403,
        6,
        408,
        4,
    ])

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font

    column_widths = {
        "A": 10, "B": 22, "C": 28, "D": 12, "E": 18, "F": 18, "G": 16,
        "H": 12, "I": 18, "J": 18,
    }
    for letter, width in column_widths.items():
        ws.column_dimensions[letter].width = width
    for index in range(11, len(POLICY_TEMPLATE_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 18

    dropdown_sheet = wb.create_sheet("Dropdowns")
    dropdown_data = {
        "A": ["TRUE", "FALSE"],
        "B": ["ATTRIBUTE", "ROLE"],
        "C": ["APPROVE", "REJECT"],
        "D": ["Weekly", "Monthly", "Quartely", "Yearly"],
        "E": ["EARNED", "FIXED", "FIXED_EARNED", "FIXED_MONTHLY_ADVANCED"],
        "F": ["MONTHLY", "NONE", "YEARLY"],
        "G": ["DAY", "MONTH", "YEAR", "YEAR_END"],
    }
    for column, values in dropdown_data.items():
        dropdown_sheet[f"{column}1"] = {
            "A": "Boolean",
            "B": "Rule_Type_1",
            "C": "Rule_Type_2",
            "D": "Grant_StartDate_Hints",
            "E": "Grant_Type",
            "F": "Grant_Frequency",
            "G": "Grant_Expiration",
        }[column]
        for row_index, value in enumerate(values, start=2):
            dropdown_sheet[f"{column}{row_index}"] = value

    validations = [
        ("=Dropdowns!$A$2:$A$3", ["H", "K", "L", "M", "N", "O", "X", "AB"]),
        ("=Dropdowns!$E$2:$E$5", ["E"]),
        ("=Dropdowns!$F$2:$F$4", ["F"]),
        ("=Dropdowns!$G$2:$G$5", ["I"]),
    ]
    for formula, columns in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(validation)
        for column in columns:
            validation.add(f"{column}2:{column}1000")

    existing_ws = wb.create_sheet(EXISTING_POLICIES_SHEET)
    existing_ws.append(POLICY_TEMPLATE_COLUMNS)
    for row in policy_rows:
        existing_ws.append([row.get(column, "") for column in POLICY_TEMPLATE_COLUMNS])

    accrual_ws = wb.create_sheet(EXISTING_ACCRUALS_SHEET)
    accrual_columns = ["id", "name", "description"]
    accrual_ws.append(accrual_columns)
    for item in accrual_rows:
        accrual_ws.append([item.get(column, "") for column in accrual_columns])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_payload(row: pd.Series) -> dict[str, Any]:
    name = _clean_text(row.get("name"))
    description = _clean_text(row.get("description")) or name
    accrual_id = _int_from_cell(row.get("accrualId"))

    if not name:
        raise ValueError("name is required")
    if accrual_id is None:
        raise ValueError("accrualId is required")

    grant_prorations = []
    for index in range(1, MAX_PRORATIONS + 1):
        start_date = _clean_text(row.get(f"startDate(DD/MM){index}"))
        end_date = _clean_text(row.get(f"sendDate(DD/MM){index}"))
        amount = _float_from_cell(row.get(f"amount{index}"))
        if not start_date and not end_date and amount is None:
            continue
        if not start_date or not end_date or amount is None:
            raise ValueError(f"grantProrations entry {index} is incomplete")
        grant_prorations.append({"startDate": start_date, "endDate": end_date, "amount": amount})

    grant_amounts = []
    for index in range(1, MAX_GRANT_AMOUNTS + 1):
        start = _int_from_cell(row.get(f"grantStart{index}"))
        end = _int_from_cell(row.get(f"grantEnd{index}"))
        maximum = row.get(f"grantMax{index}")
        amount = _float_from_cell(row.get(f"grantAmount{index}"))
        if start is None and end is None and amount is None and _clean_text(maximum) == "":
            continue
        if start is None or amount is None:
            raise ValueError(f"grantAmounts entry {index} requires grantStart and grantAmount")
        item: dict[str, Any] = {"start": start, "amount": amount}
        if end is not None:
            item["end"] = end
        if _clean_text(maximum) != "":
            item["max"] = _bool_from_cell(maximum)
        grant_amounts.append(item)

    grant_amount_rules = []
    for index in range(1, MAX_RULES + 1):
        condition = _clean_text(row.get(f"ruleCondition{index}"))
        value = _clean_text(row.get(f"ruleValue{index}"))
        if not condition and not value:
            continue
        if not condition or not value:
            raise ValueError(f"grantAmountRules entry {index} is incomplete")
        grant_amount_rules.append({"condition": condition, "value": value})

    paycodes = []
    for index in range(1, MAX_PAYCODES + 1):
        paycode_id = _int_from_cell(row.get(f"takingPaycodeID{index}"))
        amount = _float_from_cell(row.get(f"takingAmount{index}"))
        if paycode_id is None and amount is None:
            continue
        if paycode_id is None or amount is None:
            raise ValueError(f"paycodes entry {index} is incomplete")
        paycodes.append({"paycode": {"id": paycode_id}, "amount": amount})

    return {
        "name": name,
        "description": description,
        "accrual": {"id": accrual_id},
        "grantType": _clean_text(row.get("grantType")),
        "grantFrequency": _clean_text(row.get("grantFrequency")),
        "grantStartDate": _clean_text(row.get("grantStartDate")),
        "forceAvail": _bool_from_cell(row.get("forceAvail")),
        "grantExpiration": _clean_text(row.get("grantExpiration")),
        "grantExpiredAfter": _int_from_cell(row.get("grantExpiredAfter")),
        "carryoverAmountMax": _bool_from_cell(row.get("carryoverAmountMax")),
        "prioritizeCarryoverBalance": _bool_from_cell(row.get("prioritizeCarryoverBalance")),
        "carryoverEncashmentAmountMax": _bool_from_cell(row.get("carryoverEncashmentAmountMax")),
        "terminationEncashmentAmountMax": _bool_from_cell(row.get("terminationEncashmentAmountMax")),
        "manualEncashmentAmountMax": _bool_from_cell(row.get("manualEncashmentAmountMax")),
        "grantProrations": grant_prorations,
        "grantAmounts": grant_amounts,
        "grantAmountRules": grant_amount_rules,
        "paycodes": paycodes,
    }


def accrual_policies_ui() -> None:
    module_header("📊 Accrual Policies", "Create, update, delete, review, and bulk upload accrual policies")

    if not st.session_state.get("token"):
        st.error("Please login first")
        return

    host = st.session_state.HOST.rstrip("/")
    base_url = f"{host}/resource-server/api/accrual_policies"
    accrual_reference_url = f"{ACCRUAL_REFERENCE_HOST}/resource-server/api/accruals"
    headers = _auth_headers()

    policy_data, policy_error = _fetch_json(base_url, headers)
    accrual_data, accrual_error = _fetch_json(accrual_reference_url, headers)

    col1, col2, col3 = st.columns(3)
    col1.metric("Policies fetched", len(policy_data))
    col2.metric("Accrual references", len(accrual_data))
    col3.metric("Template sheets", 4)

    if policy_error:
        st.warning(f"Could not fetch existing accrual policies: {policy_error}")
    if accrual_error:
        st.warning(f"Could not fetch accrual references for template: {accrual_error}")

    st.info(
        "Use the Excel template for bulk create or update. Add an id to update an existing policy; leave id blank to create a new one."
    )

    section_header("📥 Download Upload Template")
    policy_rows = _build_existing_policy_rows(policy_data)
    accrual_rows = [
        {
            "id": item.get("id"),
            "name": item.get("name"),
            "description": item.get("description"),
        }
        for item in accrual_data
    ]
    template_bytes = _build_template_workbook(policy_rows, accrual_rows)

    st.download_button(
        "⬇️ Download Accrual Policies Template",
        data=template_bytes,
        file_name="accrual_policies_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    with st.expander("Template guide", expanded=False):
        st.markdown(
            """
            - **Sheet 1:** upload rows for create/update.
            - **Sheet 2:** flattened copy of existing accrual policies.
            - **Sheet 3:** accrual ids and names fetched from the accruals API.
            - **Sheet 4:** dropdown helper values for booleans and key enums.
            """
        )

    st.divider()
    section_header("📤 Upload Accrual Policies")

    uploaded_file = st.file_uploader("Upload Excel or CSV", type=["xlsx", "xls", "csv"])
    if uploaded_file:
        dataframe = (
            pd.read_csv(uploaded_file)
            if uploaded_file.name.endswith(".csv")
            else pd.read_excel(uploaded_file, sheet_name=0)
        ).fillna("")

        st.success(f"Rows detected: {len(dataframe)}")
        st.dataframe(dataframe, use_container_width=True, height=320)

        if st.button("🚀 Submit Accrual Policies", type="primary", use_container_width=True):
            results = []
            with st.spinner("Processing accrual policies..."):
                for index, row in dataframe.iterrows():
                    try:
                        payload = _build_payload(row)
                        policy_id = _int_from_cell(row.get("id"))
                        if policy_id is not None:
                            payload["id"] = policy_id
                            response = requests.put(f"{base_url}/{policy_id}", headers=headers, json=payload, timeout=30)
                            action = "UPDATE"
                        else:
                            response = requests.post(base_url, headers=headers, json=payload, timeout=30)
                            action = "CREATE"

                        results.append(
                            {
                                "Row": index + 1,
                                "ID": policy_id or "",
                                "Name": payload["name"],
                                "Action": action,
                                "Status": "SUCCESS" if response.status_code in (200, 201) else f"FAILED ({response.status_code})",
                                "Response": response.text[:200],
                            }
                        )
                    except Exception as exc:  # noqa: BLE001
                        results.append(
                            {
                                "Row": index + 1,
                                "ID": row.get("id", ""),
                                "Name": row.get("name", ""),
                                "Action": "ERROR",
                                "Status": str(exc),
                                "Response": "",
                            }
                        )

            result_df = pd.DataFrame(results)
            st.dataframe(result_df, use_container_width=True)
            success_count = (result_df["Status"] == "SUCCESS").sum() if not result_df.empty else 0
            st.caption(f"Processed {len(result_df)} rows. Successful rows: {success_count}.")

    st.divider()
    section_header("🗑️ Delete Accrual Policies")

    delete_ids = st.text_input("Enter policy ids (comma-separated)", placeholder="101, 102, 103")
    if st.button("Delete Accrual Policies", use_container_width=True):
        ids = [item.strip() for item in delete_ids.split(",") if item.strip().isdigit()]
        if not ids:
            st.warning("Please enter at least one valid numeric id.")
        else:
            with st.spinner("Deleting accrual policies..."):
                for policy_id in ids:
                    response = requests.delete(f"{base_url}/{policy_id}", headers=headers, timeout=30)
                    if response.status_code in (200, 204):
                        st.success(f"Deleted accrual policy {policy_id}")
                    else:
                        st.error(f"Failed to delete {policy_id}: HTTP {response.status_code}")

    st.divider()
    section_header("⬇️ Download Existing Accrual Policies")

    export_df = pd.DataFrame(policy_rows)
    if export_df.empty:
        st.caption("No accrual policies available to export right now.")
    else:
        st.dataframe(export_df.head(20), use_container_width=True, height=260)

    export_buffer = io.BytesIO()
    export_df.to_excel(export_buffer, index=False)
    export_buffer.seek(0)
    st.download_button(
        "⬇️ Download Existing Accrual Policies",
        data=export_buffer.getvalue(),
        file_name="accrual_policies_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
