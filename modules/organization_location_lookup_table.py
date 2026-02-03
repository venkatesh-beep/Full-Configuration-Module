import streamlit as st
import pandas as pd
import requests
import io
from openpyxl.styles import PatternFill, Font

# ======================================================
# HELPER: CLEAN EXCEL VALUES (REMOVE .0 ISSUE)
# ======================================================
def clean_excel_value(val):
    if pd.isna(val):
        return ""

    val_str = str(val).strip()

    # Remove trailing .0 from numbers like 11650029.0
    if val_str.endswith(".0"):
        try:
            return str(int(float(val_str)))
        except:
            pass

    # Handle pandas float values
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()

    return val_str


# ======================================================
# MAIN UI
# ======================================================
def organization_location_lookup_table_ui():
    st.header("🏢 Organization Location Lookup Table")
    st.caption("Download and upload Organization Location Lookup Table")

    BASE_URL = st.session_state.HOST.rstrip("/")
    GET_URL = BASE_URL + "/resource-server/api/organization_location_lookup_table"
    POST_URL = BASE_URL + "/resource-server/api/organization_location_lookup_table/action/"

    headers_auth = {
        "Authorization": f"Bearer {st.session_state.token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    # ==================================================
    # FETCH LOOKUP TABLE
    # ==================================================
    def fetch_lookup_table():
        r = requests.get(GET_URL, headers=headers_auth, timeout=30)
        if r.status_code != 200:
            return None, None

        raw = r.json()

        headers_meta = sorted(
            raw.get("headers", []),
            key=lambda x: x.get("sequence", 999)
        )

        data = raw.get("content") or raw.get("data") or []
        return headers_meta, data

    # ==================================================
    # DOWNLOAD EXISTING DATA
    # ==================================================
    st.subheader("⬇️ Download Existing Data")

    if st.button("⬇️ Download Existing Organization Location Data", use_container_width=True):
        with st.spinner("Preparing download..."):

            headers_meta, data = fetch_lookup_table()
            if not headers_meta:
                st.error("❌ Failed to fetch lookup data")
                return

            columns = [h["data"] for h in headers_meta]
            input_columns = [h["data"] for h in headers_meta if h.get("type") == "INPUT"]

            df = pd.DataFrame(data).reindex(columns=columns) if data else pd.DataFrame(columns=columns)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Existing_Data")
                ws = writer.book["Existing_Data"]

                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                bold_font = Font(bold=True)

                for col_idx, col_name in enumerate(columns, start=1):
                    if col_name in input_columns:
                        cell = ws.cell(row=1, column=col_idx)
                        cell.fill = red_fill
                        cell.font = bold_font

            st.download_button(
                label="⬇️ Download",
                data=output.getvalue(),
                file_name="organization_location_lookup_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.divider()

    # ==================================================
    # UPLOAD DATA WITH FIX
    # ==================================================
    st.subheader("📤 Upload Organization Location Lookup Table")

    uploaded_file = st.file_uploader(
        "Upload Excel file (same format as downloaded)",
        ["xlsx", "xls"]
    )

    if uploaded_file:
        df_upload = pd.read_excel(uploaded_file).fillna("")
        st.info(f"Rows detected: {len(df_upload)}")

        # ✅ CLEAN WHOLE DATAFRAME (fix .0 everywhere)
        df_upload = df_upload.applymap(clean_excel_value)

        if st.button("🚀 Validate & Upload", type="primary"):
            with st.spinner("Validating and uploading data..."):

                headers_meta, _ = fetch_lookup_table()
                if not headers_meta:
                    st.error("❌ Failed to fetch headers for validation")
                    return

                input_columns = [h["data"] for h in headers_meta if h.get("type") == "INPUT"]
                all_columns = [h["data"] for h in headers_meta]

                validation_errors = []
                data_rows = []

                for idx, row in df_upload.iterrows():
                    excel_row = idx + 2
                    record = {}
                    row_has_error = False

                    # INPUT VALIDATION
                    for col in input_columns:
                        value = clean_excel_value(row.get(col, ""))
                        if value == "":
                            validation_errors.append({
                                "Row": excel_row,
                                "Field": col,
                                "Error": "INPUT field cannot be empty"
                            })
                            row_has_error = True

                    if row_has_error:
                        continue

                    # BUILD DATA
                    for col in all_columns:
                        value = clean_excel_value(row.get(col, ""))
                        if value != "":
                            record[col] = value

                    if record:
                        data_rows.append(record)

                if validation_errors:
                    st.error("❌ Validation failed. Fix the errors below and re-upload.")
                    st.dataframe(pd.DataFrame(validation_errors), use_container_width=True)
                    return

                if not data_rows:
                    st.warning("No valid rows found to upload")
                    return

                payload = {
                    "action": "SAVE",
                    "table": {
                        "entityType": "ORGANIZATION_LOCATION",
                        "headers": headers_meta,
                        "data": data_rows
                    }
                }

                r = requests.post(POST_URL, headers=headers_auth, json=payload, timeout=60)

                if r.status_code in (200, 201):
                    st.success("✅ Organization Location Lookup Table updated successfully")
                else:
                    st.error("❌ Upload failed")
                    st.code(f"{r.status_code}\n{r.text}")
