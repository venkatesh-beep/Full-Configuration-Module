import io
from typing import Any

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from modules.ui_helpers import module_header, section_header


BOOLEAN_OPTIONS = ["TRUE", "FALSE"]
APPROVER_TYPE_OPTIONS = ["ATTRIBUTE", "ROLE"]
TAT_ACTION_OPTIONS = ["APPROVE", "REJECT"]
PERIOD_OPTIONS = ["Weekly", "Monthly", "Quartely", "Yearly"]
TEMPLATE_COLUMNS = [
    "id",
    "name",
    "description",
    "attendanceRegularizationTypeId",
    "numOfPastSignedOfPeriods",
    "approvalLevel",
    "approverType",
    "remarks",
    "considerSignOff",
    "enableForEmployee",
    "remarksRequired",
    "usageApplicable",
    "usageCount",
    "referenceDate",
    "period",
    "level1",
    "approver1",
    "sendNotification1",
    "reminderNotificationDurations1",
    "tatDuration1",
    "tatAction1",
    "sendEmployeeNotification1",
    "pushNotification1",
    "level2",
    "approver2",
    "sendNotification2",
    "reminderNotificationDurations2",
    "tatDuration2",
    "tatAction2",
    "sendEmployeeNotification2",
    "pushNotification2",
]


def _safe_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None



def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()



def _to_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value

    text = str(value).strip().lower()
    if text == "":
        return None
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None



def _format_bool(value: Any) -> str:
    parsed = _to_bool(value)
    if parsed is None:
        return ""
    return "TRUE" if parsed else "FALSE"



def _json_headers(token: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }



def _fetch_json(url: str, headers: dict[str, str]) -> tuple[list[dict[str, Any]], str | None]:
    try:
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code != 200:
            return [], f"GET {url} failed with status {response.status_code}: {response.text[:300]}"
        data = response.json()
        if isinstance(data, list):
            return data, None
        return [], f"GET {url} returned unexpected response format"
    except requests.RequestException as exc:
        return [], str(exc)



def _build_template_workbook(type_rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload_Template"
    ws.append(TEMPLATE_COLUMNS)

    dropdowns = wb.create_sheet("Dropdowns")
    dropdowns.append(["Boolean", "ApproverType", "TatAction", "Period"])
    max_len = max(
        len(BOOLEAN_OPTIONS),
        len(APPROVER_TYPE_OPTIONS),
        len(TAT_ACTION_OPTIONS),
        len(PERIOD_OPTIONS),
    )
    for idx in range(max_len):
        dropdowns.append(
            [
                BOOLEAN_OPTIONS[idx] if idx < len(BOOLEAN_OPTIONS) else "",
                APPROVER_TYPE_OPTIONS[idx] if idx < len(APPROVER_TYPE_OPTIONS) else "",
                TAT_ACTION_OPTIONS[idx] if idx < len(TAT_ACTION_OPTIONS) else "",
                PERIOD_OPTIONS[idx] if idx < len(PERIOD_OPTIONS) else "",
            ]
        )

    types_sheet = wb.create_sheet("Attendance_Regularization_Types")
    types_columns = ["id", "name", "description"]
    types_sheet.append(types_columns)
    for row in type_rows:
        types_sheet.append([
            row.get("id"),
            row.get("name"),
            row.get("description"),
        ])

    validations = {
        "G2:G1000": "=Dropdowns!$B$2:$B$3",
        "I2:I1000": "=Dropdowns!$A$2:$A$3",
        "J2:J1000": "=Dropdowns!$A$2:$A$3",
        "K2:K1000": "=Dropdowns!$A$2:$A$3",
        "L2:L1000": "=Dropdowns!$A$2:$A$3",
        "S2:S1000": "=Dropdowns!$A$2:$A$3",
        "U2:U1000": "=Dropdowns!$C$2:$C$3",
        "V2:V1000": "=Dropdowns!$A$2:$A$3",
        "W2:W1000": "=Dropdowns!$A$2:$A$3",
        "Z2:Z1000": "=Dropdowns!$A$2:$A$3",
        "AB2:AB1000": "=Dropdowns!$C$2:$C$3",
        "AC2:AC1000": "=Dropdowns!$A$2:$A$3",
        "AD2:AD1000": "=Dropdowns!$A$2:$A$3",
        "O2:O1000": "=Dropdowns!$D$2:$D$5",
    }

    for cell_range, formula in validations.items():
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(cell_range)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def _extract_approval_levels(policy: dict[str, Any]) -> list[dict[str, Any]]:
    levels = policy.get("approvalLevels")
    if isinstance(levels, list) and levels:
        return levels

    extracted = []
    for index in (1, 2):
        approver = policy.get(f"approver{index}")
        if approver is None and policy.get(f"level{index}") is None:
            continue
        extracted.append(
            {
                "level": policy.get(f"level{index}"),
                "approver": approver,
                "sendNotification": policy.get(f"sendNotification{index}"),
                "reminderNotificationDurations": policy.get(f"reminderNotificationDurations{index}"),
                "tatDuration": policy.get(f"tatDuration{index}"),
                "tatAction": policy.get(f"tatAction{index}"),
                "sendEmployeeNotification": policy.get(f"sendEmployeeNotification{index}"),
                "pushNotification": policy.get(f"pushNotification{index}"),
            }
        )
    return extracted



def _policy_to_row(policy: dict[str, Any]) -> dict[str, Any]:
    attendance_type = policy.get("attendanceRegularizationType") or {}
    row = {
        "id": policy.get("id"),
        "name": policy.get("name"),
        "description": policy.get("description"),
        "attendanceRegularizationTypeId": attendance_type.get("id") or policy.get("attendanceRegularizationTypeId"),
        "numOfPastSignedOfPeriods": policy.get("numOfPastSignedOfPeriods"),
        "approvalLevel": policy.get("approvalLevel"),
        "approverType": policy.get("approverType"),
        "remarks": policy.get("remarks"),
        "considerSignOff": _format_bool(policy.get("considerSignOff")),
        "enableForEmployee": _format_bool(policy.get("enableForEmployee")),
        "remarksRequired": _format_bool(policy.get("remarksRequired")),
        "usageApplicable": _format_bool(policy.get("usageApplicable")),
        "usageCount": policy.get("usageCount"),
        "referenceDate": policy.get("referenceDate"),
        "period": policy.get("period"),
    }

    for index, level in enumerate(_extract_approval_levels(policy), start=1):
        row[f"level{index}"] = level.get("level")
        row[f"approver{index}"] = level.get("approver")
        row[f"sendNotification{index}"] = _format_bool(level.get("sendNotification"))
        row[f"reminderNotificationDurations{index}"] = level.get("reminderNotificationDurations")
        row[f"tatDuration{index}"] = level.get("tatDuration")
        row[f"tatAction{index}"] = level.get("tatAction")
        row[f"sendEmployeeNotification{index}"] = _format_bool(level.get("sendEmployeeNotification"))
        row[f"pushNotification{index}"] = _format_bool(level.get("pushNotification"))

    return row



def _row_to_payload(row: pd.Series) -> tuple[dict[str, Any], list[str]]:
    errors: list[str] = []
    name = _normalize_text(row.get("name"))
    description = _normalize_text(row.get("description")) or name
    attendance_type_id = _safe_int(row.get("attendanceRegularizationTypeId"))

    if not name:
        errors.append("name is required")
    if attendance_type_id is None:
        errors.append("attendanceRegularizationTypeId is required")

    payload: dict[str, Any] = {
        "name": name,
        "description": description,
        "attendanceRegularizationType": {"id": attendance_type_id} if attendance_type_id is not None else None,
        "numOfPastSignedOfPeriods": _safe_int(row.get("numOfPastSignedOfPeriods")),
        "approvalLevel": _safe_int(row.get("approvalLevel")),
        "approverType": _normalize_text(row.get("approverType")),
        "remarks": _normalize_text(row.get("remarks")),
        "considerSignOff": _to_bool(row.get("considerSignOff")),
        "enableForEmployee": _to_bool(row.get("enableForEmployee")),
        "remarksRequired": _to_bool(row.get("remarksRequired")),
        "usageApplicable": _to_bool(row.get("usageApplicable")),
        "usageCount": _safe_int(row.get("usageCount")),
        "referenceDate": _normalize_text(row.get("referenceDate")),
        "period": _normalize_text(row.get("period")),
    }

    approval_levels = []
    for index in (1, 2):
        level_value = _safe_int(row.get(f"level{index}"))
        approver_value = _normalize_text(row.get(f"approver{index}"))
        reminder_value = _normalize_text(row.get(f"reminderNotificationDurations{index}"))
        tat_duration = _normalize_text(row.get(f"tatDuration{index}"))
        tat_action = _normalize_text(row.get(f"tatAction{index}"))
        send_notification = _to_bool(row.get(f"sendNotification{index}"))
        send_employee_notification = _to_bool(row.get(f"sendEmployeeNotification{index}"))
        push_notification = _to_bool(row.get(f"pushNotification{index}"))

        if any(
            value not in (None, "")
            for value in [
                level_value,
                approver_value,
                reminder_value,
                tat_duration,
                tat_action,
                send_notification,
                send_employee_notification,
                push_notification,
            ]
        ):
            approval_levels.append(
                {
                    "level": level_value,
                    "approver": approver_value,
                    "sendNotification": send_notification,
                    "reminderNotificationDurations": reminder_value,
                    "tatDuration": tat_duration,
                    "tatAction": tat_action,
                    "sendEmployeeNotification": send_employee_notification,
                    "pushNotification": push_notification,
                }
            )

    if approval_levels:
        payload["approvalLevels"] = approval_levels

    payload = {key: value for key, value in payload.items() if value not in (None, "")}
    return payload, errors



def regularization_policies_ui() -> None:
    module_header(
        "📊 Regularization Policies",
        "Create, update, delete, import and export regularization policies",
    )

    token = st.session_state.get("token")
    if not token:
        st.error("Please login first")
        return

    host = st.session_state.HOST.rstrip("/")
    base_url = f"{host}/resource-server/api/regularization_policies"
    attendance_types_url = f"{host}/resource-server/api/attendance_regularization_types"
    headers = _json_headers(token)

    section_header("📥 Download Upload Template")

    attendance_types, types_error = _fetch_json(attendance_types_url, headers)
    workbook_bytes = _build_template_workbook(attendance_types)
    st.download_button(
        "⬇️ Download Template",
        data=workbook_bytes,
        file_name="regularization_policies_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    if types_error:
        st.warning(f"Template Sheet 2 could not load attendance regularization types dynamically: {types_error}")

    st.divider()

    section_header("📤 Upload Regularization Policies")
    uploaded_file = st.file_uploader("Upload CSV or Excel file", ["csv", "xlsx", "xls"])

    if uploaded_file:
        df = (
            pd.read_csv(uploaded_file)
            if uploaded_file.name.endswith(".csv")
            else pd.read_excel(uploaded_file)
        )
        df = df.fillna("")
        st.success(f"File loaded successfully — {len(df)} rows")
        st.dataframe(df, use_container_width=True)

        if st.button("🚀 Process Upload", type="primary", use_container_width=True):
            results = []
            with st.spinner("⏳ Processing regularization policies..."):
                for index, row in df.iterrows():
                    payload, errors = _row_to_payload(row)
                    row_id = _safe_int(row.get("id"))
                    name = _normalize_text(row.get("name"))

                    if errors:
                        results.append(
                            {
                                "Row": index + 1,
                                "Name": name,
                                "Action": "Skipped",
                                "Status": "Failed",
                                "HTTP Status": "",
                                "Message": ", ".join(errors),
                            }
                        )
                        continue

                    try:
                        if row_id is not None:
                            payload["id"] = row_id
                            response = requests.put(
                                f"{base_url}/{row_id}",
                                headers=headers,
                                json=payload,
                                timeout=30,
                            )
                            action = "Update"
                        else:
                            response = requests.post(
                                base_url,
                                headers=headers,
                                json=payload,
                                timeout=30,
                            )
                            action = "Create"

                        results.append(
                            {
                                "Row": index + 1,
                                "Name": name,
                                "Action": action,
                                "Status": "Success" if response.status_code in (200, 201) else "Failed",
                                "HTTP Status": response.status_code,
                                "Message": response.text[:250],
                            }
                        )
                    except requests.RequestException as exc:
                        results.append(
                            {
                                "Row": index + 1,
                                "Name": name,
                                "Action": "Error",
                                "Status": "Failed",
                                "HTTP Status": "",
                                "Message": str(exc),
                            }
                        )

            section_header("📊 Upload Result")
            st.dataframe(pd.DataFrame(results), use_container_width=True)

    st.divider()

    section_header("🗑️ Delete Regularization Policies")
    ids_input = st.text_input("Enter Regularization Policy IDs (comma-separated)", placeholder="101,102")
    if st.button("Delete Regularization Policies", use_container_width=True):
        ids = [item.strip() for item in ids_input.split(",") if item.strip().isdigit()]
        if not ids:
            st.warning("Please enter at least one numeric id")
        else:
            for policy_id in ids:
                try:
                    response = requests.delete(f"{base_url}/{policy_id}", headers=headers, timeout=30)
                    if response.status_code in (200, 204):
                        st.success(f"Deleted {policy_id}")
                    else:
                        st.error(f"Failed to delete {policy_id}: {response.status_code} {response.text[:200]}")
                except requests.RequestException as exc:
                    st.error(f"Failed to delete {policy_id}: {exc}")

    st.divider()

    section_header("⬇️ Download Existing Regularization Policies")
    policies, policies_error = _fetch_json(base_url, headers)
    if policies_error:
        st.error(f"Failed to fetch regularization policies: {policies_error}")
        return

    export_df = pd.DataFrame([_policy_to_row(policy) for policy in policies], columns=TEMPLATE_COLUMNS)
    export_output = io.BytesIO()
    with pd.ExcelWriter(export_output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Regularization_Policies")
    export_output.seek(0)

    st.download_button(
        "⬇️ Download Existing Regularization Policies",
        data=export_output.getvalue(),
        file_name="regularization_policies_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
